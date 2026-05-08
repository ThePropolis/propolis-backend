"""
One-time CLI import for canonical inventory.

Reads Documentations/Properties Sheet/Propolis Properties Overview.xlsx and
upserts inv_buildings / inv_units / inv_rooms idempotently. Re-running the
script only inserts rows that don't already exist (matched by name within
their parent), so any owner edits made in the website are preserved.

Usage:
    cd Backend
    python3 scripts/import_inventory.py
"""
from __future__ import annotations

import os
import re
import sys
from itertools import takewhile
from pathlib import Path

# Make sibling modules importable when run as a script
SCRIPT_DIR = Path(__file__).resolve().parent
BACKEND_DIR = SCRIPT_DIR.parent
sys.path.insert(0, str(BACKEND_DIR))

from dotenv import load_dotenv
import openpyxl

load_dotenv(BACKEND_DIR / ".env")

from database import supabase  # noqa: E402

XLSX_PATH = (
    BACKEND_DIR.parent
    / "Documentations"
    / "Properties Sheet"
    / "Propolis Properties Overview.xlsx"
)

SHEETS_TO_SKIP = {"Anatoly", "P&L"}

# Column headers (case-insensitive) that are treated as boolean amenity flags.
# When the row has "Yes" in the column, the header (canonicalized) is appended
# to that room's amenities[]. Anything else is ignored for the amenities array.
AMENITY_FLAG_HEADERS = {
    "dishwasher",
    "mini fridge",
    "plugin cooktop",
    "washer",
    "dryer",
    "ac",
    "air conditioner",
    "wifi",
    "furnished",
    "balcony",  # only emitted as amenity when the cell is Yes; the descriptive
                # version (e.g. "Personal Balcony") goes into the balcony column
}

# Headers we want to capture as their own structured column (not as amenity flags)
STRUCTURED_HEADERS = {
    "length": "length",
    "strategy": "strategy",
    "bed size": "bed_size",
    "bathroom": "bathroom",
    "ceiling height": "ceiling_height",
    "ceiling heigh": "ceiling_height",  # typo in Plum sheet
    "balcony": "balcony",  # also amenity-flag, handled specially below
    "room type name": "room_type_name",
}

# Headers that identify the unit/room cell. The first column header for room
# rows is usually one of these.
UNIT_LABEL_HEADERS = {"unit", "apartment", "bedroom"}

# Per-room money columns. Many sheets use slightly different headings, so we
# map a wide set of synonyms onto our canonical columns.
RENT_HEADERS = {
    "actual_rent":   ["actual rent", "rent(current)", "rent (current)", "current rent"],
    "base_rent":     ["base rent", "monthly", "forecasted montly", "forecasted monthly"],
    "market_rent":   ["market rent"],
    "revenue_year":  ["revenue year", "2024 revenue", "revenue 2024", "annual revenue"],
    "revenue_month": ["revenue month", "monthly revenue"],
    "extras":        ["extras"],
    "adjustment":    ["adjustment"],
    "actual_rent_with_util":     ["actual rent + util", "actual rent + utilities", "market rent + util", "market rent + utilities"],
    "pessimistic_rent":          ["pesimistic approach with utilities", "pessimistic approach with utilities", "pessimistic rent"],
    "concession_rent":           ["concession rent (13 mo)", "concession rent"],
    "concession_rent_with_util": ["concession rent + util", "concession rent + utilities"],
    "stake_5_cashback":          ["stake 5% cashback", "stake 5 cashback"],
    "stake_8_cashback":          ["stake 8% cashback", "stake 8 cashback"],
    "revenue_per_apartment":     ["revenue per apartment shared", "revenue per apartment"],
    "unit_type":                 ["unit type"],
}
ADA_HEADERS = ["ada units", "ada units ", "ada unit", "ada"]
LISTING_DATE_HEADERS = ["beds/bath", "beds / bath"]


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("$", "").replace(",", "").replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def find_money_col(headers: dict, key: str) -> int | None:
    for cand in RENT_HEADERS[key]:
        if cand in headers:
            return headers[cand]
    return None


def normalize_length(value):
    if value is None:
        return None
    v = str(value).strip().lower()
    if not v:
        return None
    if v.startswith("short"):
        return "STR"
    if v.startswith("long"):
        return "LTR"
    return None


def normalize_strategy(value):
    if value is None:
        return None
    v = str(value).strip()
    if not v:
        return None
    low = v.lower()
    if "coliv" in low:
        return "Coliving"
    if "entire" in low:
        return "Entire Apt"
    return v  # leave anything else as-is


def normalize_yes_no_to_bool(value):
    if value is None:
        return None
    v = str(value).strip().lower()
    if v in ("yes", "true", "y"):
        return True
    if v in ("no", "false", "n"):
        return False
    return None


def to_camel(s: str) -> str:
    """'Mini Fridge' -> 'MiniFridge'"""
    return "".join(w.capitalize() for w in re.split(r"[\s_-]+", s.strip()) if w)


def parse_room_identifier(value):
    """
    Given a cell value like '11A', '11', 21.0, returns the room name as str.
    Numbers are coerced to ints when whole.
    """
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def derive_unit_name(room_name: str) -> str:
    """11A -> 11; 21S -> 21; 11 -> 11"""
    digits = "".join(takewhile(str.isdigit, room_name))
    return digits or room_name


def parse_building_meta(ws):
    """Header rows 1-6: extract building-level fields."""
    def cell(row, col):
        return ws.cell(row=row, column=col).value

    # Owner LLC: scan row 1 for an 'Owner' label, take the next cell to the right.
    owner_llc = None
    for c in range(1, ws.max_column + 1):
        v = cell(1, c)
        if v and str(v).strip().lower() == "owner":
            owner_llc = cell(1, c + 1)
            break

    return {
        "full_name": cell(1, 2),  # B1
        "owner_llc": owner_llc,
        "address": cell(2, 2),
        "units_count": _to_int(cell(3, 2)),
        "beds_count": _to_int(cell(4, 2)),
        "floors": _to_int(cell(5, 2)),
        "has_elevator": normalize_yes_no_to_bool(cell(6, 2)),
    }


def _to_int(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    try:
        return int(str(v).strip())
    except (ValueError, TypeError):
        return None


def parse_headers(ws):
    """Row 8 = headers. Returns dict mapping lower-cased header → 1-based column."""
    headers: dict = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=8, column=c).value
        if not v:
            continue
        key = str(v).strip().lower()
        if key and key not in headers:  # first occurrence wins
            headers[key] = c
    return headers


def find_first_col(headers: dict, candidates) -> int | None:
    for cand in candidates:
        if cand in headers:
            return headers[cand]
    return None


def parse_rooms(ws, headers: dict):
    """
    Yields (unit_name, room_payload) tuples for each data row.
    Auto-detects the room-id column (Unit / Bedroom) and the apartment column
    (Apartment) for split layouts like Olive.
    """
    apartment_col = headers.get("apartment")
    room_col = find_first_col(headers, ["bedroom", "unit"])
    if room_col is None:
        # Last resort — first column of row 8
        room_col = 1

    length_col = headers.get("length")
    strategy_col = headers.get("strategy")
    bed_col = headers.get("bed size")
    bath_col = headers.get("bathroom")
    ceil_col = headers.get("ceiling height") or headers.get("ceiling heigh")
    balcony_col = headers.get("balcony")
    room_type_col = headers.get("room type name")
    actual_rent_col  = find_money_col(headers, "actual_rent")
    base_rent_col    = find_money_col(headers, "base_rent")
    market_rent_col  = find_money_col(headers, "market_rent")
    revenue_year_col = find_money_col(headers, "revenue_year")
    revenue_mo_col   = find_money_col(headers, "revenue_month")
    extras_col       = find_money_col(headers, "extras")
    adjustment_col   = find_money_col(headers, "adjustment")
    arwu_col         = find_money_col(headers, "actual_rent_with_util")
    pessimistic_col  = find_money_col(headers, "pessimistic_rent")
    concession_col   = find_money_col(headers, "concession_rent")
    concession_u_col = find_money_col(headers, "concession_rent_with_util")
    stake5_col       = find_money_col(headers, "stake_5_cashback")
    stake8_col       = find_money_col(headers, "stake_8_cashback")
    rev_per_apt_col  = find_money_col(headers, "revenue_per_apartment")
    unit_type_col    = find_money_col(headers, "unit_type")
    ada_col = next((headers[h] for h in ADA_HEADERS if h in headers), None)
    listing_date_col = next((headers[h] for h in LISTING_DATE_HEADERS if h in headers), None)

    # Pre-compute (column, label) for amenity-flag columns
    amenity_flag_cols = []
    for header_text, col_idx in headers.items():
        if header_text in AMENITY_FLAG_HEADERS:
            amenity_flag_cols.append((col_idx, header_text))

    for row_idx in range(9, ws.max_row + 1):
        room_cell = ws.cell(row=row_idx, column=room_col).value
        if room_cell is None:
            # First empty row signals end of the unit table. Some sheets
            # (e.g. Olive) have unrelated summary tables further down; stop
            # parsing entirely so we don't pull garbage into inv_rooms.
            break
        room_name = parse_room_identifier(room_cell)
        if not room_name:
            break

        # Decide unit_name
        if apartment_col:
            apt_val = ws.cell(row=row_idx, column=apartment_col).value
            apt_str = parse_room_identifier(apt_val) if apt_val is not None else None
            unit_name = apt_str or derive_unit_name(room_name)
        else:
            unit_name = derive_unit_name(room_name)

        # Real apartment ids start with a digit (11, 12, 21A, 31S, etc.).
        # Anything else is a header row from a secondary table — stop here.
        if not unit_name or not unit_name[0].isdigit():
            break
        # Same sanity check for the room id
        if not re.match(r"^[\dA-Za-z]+$", room_name) or not room_name[0].isdigit():
            continue

        # Structured fields
        def cell_val(col):
            if col is None:
                return None
            v = ws.cell(row=row_idx, column=col).value
            if v is None:
                return None
            s = str(v).strip()
            return s or None

        length = normalize_length(cell_val(length_col)) if length_col else None
        strategy = normalize_strategy(cell_val(strategy_col)) if strategy_col else None

        # Amenities flag columns
        amenities: list[str] = []
        for col_idx, header_text in amenity_flag_cols:
            # 'balcony' descriptive text (e.g. "Personal Balcony") should NOT
            # contribute to the amenity list — only Yes/No flags do.
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            text = str(v).strip()
            if text.lower() in ("yes", "true", "y"):
                amenities.append(to_camel(header_text))

        def money(col):
            return _to_float(ws.cell(row=row_idx, column=col).value) if col else None

        ada_raw = ws.cell(row=row_idx, column=ada_col).value if ada_col else None
        is_ada = normalize_yes_no_to_bool(ada_raw) if ada_raw is not None else None

        # listing_date — coerce datetime-like cells to ISO date string
        listing_date = None
        if listing_date_col:
            v = ws.cell(row=row_idx, column=listing_date_col).value
            if hasattr(v, "isoformat"):
                listing_date = v.date().isoformat() if hasattr(v, "date") else v.isoformat()

        room_payload = {
            "name": room_name,
            "length": length,
            "strategy": strategy,
            "bed_size": cell_val(bed_col),
            "bathroom": cell_val(bath_col),
            "ceiling_height": cell_val(ceil_col),
            "balcony": cell_val(balcony_col),
            "room_type_name": cell_val(room_type_col),
            "amenities": amenities,
            "actual_rent":   money(actual_rent_col),
            "base_rent":     money(base_rent_col),
            "market_rent":   money(market_rent_col),
            "revenue_year":  money(revenue_year_col),
            "revenue_month": money(revenue_mo_col),
            "is_ada":        is_ada,
            "extras":        cell_val(extras_col),
            "adjustment":                money(adjustment_col),
            "actual_rent_with_util":     money(arwu_col),
            "pessimistic_rent":          money(pessimistic_col),
            "concession_rent":           money(concession_col),
            "concession_rent_with_util": money(concession_u_col),
            "stake_5_cashback":          money(stake5_col),
            "stake_8_cashback":          money(stake8_col),
            "revenue_per_apartment":     money(rev_per_apt_col),
            "unit_type":                 cell_val(unit_type_col),
            "listing_date":              listing_date,
        }
        yield unit_name, room_payload


def upsert_building(name: str, meta: dict) -> tuple[str, bool]:
    """Returns (id, was_created)."""
    existing = (
        supabase.table("inv_buildings")
        .select("id")
        .eq("name", name)
        .maybe_single()
        .execute()
    )
    if getattr(existing, "data", None):
        return existing.data["id"], False
    payload = {"name": name, **{k: v for k, v in meta.items() if v is not None}}
    resp = supabase.table("inv_buildings").insert(payload).execute()
    return resp.data[0]["id"], True


def upsert_unit(building_id: str, name: str) -> tuple[str, bool]:
    existing = (
        supabase.table("inv_units")
        .select("id")
        .eq("building_id", building_id)
        .eq("name", name)
        .maybe_single()
        .execute()
    )
    if getattr(existing, "data", None):
        return existing.data["id"], False
    resp = (
        supabase.table("inv_units")
        .insert({"building_id": building_id, "name": name})
        .execute()
    )
    return resp.data[0]["id"], True


FILLABLE_ROOM_FIELDS = (
    "length", "strategy", "bed_size", "bathroom", "ceiling_height",
    "balcony", "room_type_name", "actual_rent", "base_rent", "market_rent",
    "revenue_year", "revenue_month", "is_ada", "extras",
    "adjustment", "actual_rent_with_util", "pessimistic_rent",
    "concession_rent", "concession_rent_with_util", "stake_5_cashback",
    "stake_8_cashback", "revenue_per_apartment", "unit_type", "listing_date",
)


def upsert_room(unit_id: str, payload: dict) -> tuple[bool, bool]:
    """Returns (created, backfilled).

    On insert: fills everything from the XLSX.
    On existing row: only fills fields that are currently NULL in the DB —
    so owner edits are never overwritten, but newly-imported financial
    columns are populated on rooms imported before financials existed.
    """
    existing = (
        supabase.table("inv_rooms")
        .select("id," + ",".join(FILLABLE_ROOM_FIELDS))
        .eq("unit_id", unit_id)
        .eq("name", payload["name"])
        .maybe_single()
        .execute()
    )
    row = getattr(existing, "data", None)
    if not row:
        body = {"unit_id": unit_id, **{k: v for k, v in payload.items() if v is not None}}
        if "amenities" not in body:
            body["amenities"] = []
        supabase.table("inv_rooms").insert(body).execute()
        return True, False

    # Backfill: set columns that are NULL in the DB and have a non-null XLSX value
    updates = {}
    for field in FILLABLE_ROOM_FIELDS:
        if row.get(field) is None and payload.get(field) is not None:
            updates[field] = payload[field]
    if updates:
        supabase.table("inv_rooms").update(updates).eq("id", row["id"]).execute()
        return False, True
    return False, False


MONTH_NAME_TO_NUM = {
    "jan": 1, "january": 1,  "feb": 2, "february": 2,
    "mar": 3, "march": 3,    "apr": 4, "april": 4,
    "may": 5,                "jun": 6, "june": 6,
    "jul": 7, "july": 7,     "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


def parse_monthly_performance(ws, building_id: str) -> int:
    """
    Some sheets (Olive in particular) have a 'Top-Line Performance' table at
    the bottom: Month / Occupancy / ADR / RevPAR. Find it and import.

    Returns the number of monthly rows inserted.
    """
    # Find the row containing 'Month' as a header somewhere below the main table.
    header_row = None
    year = None
    title_re = re.compile(r"(20\d{2})")
    for r in range(20, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).strip().lower()
            if s == "month":
                header_row = r
                # Year is usually in the title row right above
                for prev in range(max(1, r - 3), r):
                    title_text = " ".join(
                        str(ws.cell(row=prev, column=cc).value or "")
                        for cc in range(1, ws.max_column + 1)
                    )
                    m = title_re.search(title_text)
                    if m:
                        year = int(m.group(1))
                        break
                break
        if header_row:
            break

    if not header_row:
        return 0

    # Map the column indices for the perf table. Row=header_row contains:
    # Month, Occupancy (%), ADR ($), RevPAR ($), [Revenue?]
    perf_cols = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if not v:
            continue
        key = str(v).strip().lower()
        if key.startswith("month"):
            perf_cols["month"] = c
        elif "occupancy" in key:
            perf_cols["occupancy_pct"] = c
        elif key.startswith("adr"):
            perf_cols["adr"] = c
        elif key.startswith("revpar"):
            perf_cols["revpar"] = c
        elif key.startswith("revenue"):
            perf_cols["revenue"] = c

    if "month" not in perf_cols:
        return 0

    inserted = 0
    for r in range(header_row + 1, ws.max_row + 1):
        month_cell = ws.cell(row=r, column=perf_cols["month"]).value
        if month_cell is None:
            continue  # totals row often has empty Month
        month_str = str(month_cell).strip().lower()
        # Skip total rows / non-month strings
        month_num = MONTH_NAME_TO_NUM.get(month_str[:3]) or MONTH_NAME_TO_NUM.get(month_str)
        if not month_num:
            continue
        if not year:
            year = 2025  # safe default if title didn't include one

        occ = _to_float(ws.cell(row=r, column=perf_cols.get("occupancy_pct", 0)).value) if perf_cols.get("occupancy_pct") else None
        # If the occupancy value is fractional (e.g. 0.78), convert to %
        if occ is not None and occ <= 1.5:
            occ = round(occ * 100, 2)
        adr = _to_float(ws.cell(row=r, column=perf_cols.get("adr", 0)).value) if perf_cols.get("adr") else None
        revpar = _to_float(ws.cell(row=r, column=perf_cols.get("revpar", 0)).value) if perf_cols.get("revpar") else None
        revenue = _to_float(ws.cell(row=r, column=perf_cols.get("revenue", 0)).value) if perf_cols.get("revenue") else None

        # Skip if it already exists
        existing = (
            supabase.table("inv_monthly_performance")
            .select("id")
            .eq("building_id", building_id)
            .eq("period_year", year)
            .eq("period_month", month_num)
            .maybe_single()
            .execute()
        )
        if getattr(existing, "data", None):
            continue
        supabase.table("inv_monthly_performance").insert({
            "building_id": building_id,
            "period_year": year,
            "period_month": month_num,
            "occupancy_pct": occ,
            "adr": adr,
            "revpar": revpar,
            "revenue": revenue,
        }).execute()
        inserted += 1
    return inserted


def import_sheet(ws) -> tuple[int, int, int, int, int]:
    """Returns (created_buildings, created_units, created_rooms, backfilled_rooms, perf_rows)."""
    name = ws.title.strip()
    meta = parse_building_meta(ws)
    headers = parse_headers(ws)

    building_id, b_created = upsert_building(name, meta)

    seen_units: dict[str, str] = {}
    units_created = 0
    rooms_created = 0
    rooms_backfilled = 0

    for unit_name, room_payload in parse_rooms(ws, headers):
        if unit_name not in seen_units:
            uid, u_created = upsert_unit(building_id, unit_name)
            seen_units[unit_name] = uid
            if u_created:
                units_created += 1
        room_payload_clean = {k: v for k, v in room_payload.items() if v is not None or k == "amenities"}
        created, backfilled = upsert_room(seen_units[unit_name], room_payload_clean)
        if created:
            rooms_created += 1
        elif backfilled:
            rooms_backfilled += 1

    perf_rows = parse_monthly_performance(ws, building_id)

    return (1 if b_created else 0), units_created, rooms_created, rooms_backfilled, perf_rows


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: XLSX not found at {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)

    print(f"Loading {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    total_b = total_u = total_r = total_bf = total_perf = 0
    for sheet_name in wb.sheetnames:
        if sheet_name in SHEETS_TO_SKIP:
            continue
        ws = wb[sheet_name]
        try:
            b, u, r, bf, perf = import_sheet(ws)
        except Exception as e:
            print(f"  ✗ {sheet_name}: {type(e).__name__}: {e}")
            continue
        total_b += b
        total_u += u
        total_r += r
        total_bf += bf
        total_perf += perf
        bits = [f"+{b} bldg", f"+{u} units", f"+{r} rooms"]
        if bf:
            bits.append(f"~{bf} room$ filled")
        if perf:
            bits.append(f"+{perf} months")
        print(f"  ✓ {sheet_name}: " + ", ".join(bits))

    print()
    print(
        f"Done. +{total_b} buildings · +{total_u} units · +{total_r} rooms · "
        f"~{total_bf} rooms with new financials · +{total_perf} monthly perf rows."
    )
    print("Re-run anytime — it's idempotent.")


if __name__ == "__main__":
    main()
